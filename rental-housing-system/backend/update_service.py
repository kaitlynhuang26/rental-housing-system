from __future__ import annotations

import json
import os
import shutil
from calendar import monthrange
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.worksheet.worksheet import Worksheet

from .excel_service import (
    DEFAULT_EXCEL_PATH,
    ExcelDataError,
    _clean_text,
    _normalize_column,
    _normalize_payment_method,
    _normalize_record_status,
    _normalize_room_status,
    _normalize_yes_no,
    _to_int,
    _to_money,
)
from .models import ChangePreview, OperationResponse


AUDIT_COLUMNS = [
    "timestamp",
    "action_type",
    "room_id",
    "sheet_name",
    "row_index",
    "old_value",
    "new_value",
    "user_message",
    "status",
]

PAYMENT_EXTRA_COLUMNS = ["record_status", "source", "notes"]


class RentalUpdateService:
    def __init__(self, excel_path: str | Path | None = None, today: date | None = None):
        configured_path = excel_path or os.getenv("EXCEL_FILE_PATH") or DEFAULT_EXCEL_PATH
        self.excel_path = Path(configured_path).expanduser().resolve()
        self.today = today or date.today()
        self.backup_dir = (
            self.excel_path.parent.parent / "backups"
            if self.excel_path.parent.name == "data"
            else self.excel_path.parent / "backups"
        )

    def auto_rollover_rental_periods(
        self,
        *,
        preview: bool = True,
        excluded_room_ids: list[int] | None = None,
        included_room_ids: list[int] | None = None,
    ) -> OperationResponse:
        workbook, sheets = self._load_update_workbook()
        payments_ws = sheets["payments"]
        payment_headers = self._headers(payments_ws)
        room_by_id = self._room_by_id(sheets["room_record"])
        payment_rows = self._payment_rows(payments_ws, payment_headers)
        latest_by_room = self._latest_active_payment_by_room(payment_rows)

        rows_to_create: list[dict[str, Any]] = []
        skipped: list[str] = []
        warnings: list[str] = []
        excluded = set(excluded_room_ids or [])
        included = set(included_room_ids or [])

        for room_id, latest in sorted(latest_by_room.items()):
            if included and room_id not in included:
                continue
            if room_id in excluded:
                skipped.append(f"Room {room_id} skipped because it was excluded by the user.")
                continue
            rent_end_date = self._parse_workbook_date(latest.get("rent_end_date"))
            rent_start_date = self._parse_workbook_date(latest.get("rent_start_date"))
            room_status = _normalize_room_status(latest.get("room_status"))
            rent_required = _normalize_yes_no(room_by_id.get(room_id, {}).get("rent_required"))

            if room_status != "Occupied":
                skipped.append(f"Room {room_id} skipped because latest status is {room_status}.")
                continue
            if rent_required == "N":
                skipped.append(f"Room {room_id} skipped because rent_required is N.")
                continue
            if rent_start_date is None or rent_end_date is None:
                warnings.append(f"Room {room_id} skipped because latest rent dates are invalid.")
                continue
            if rent_end_date < rent_start_date:
                warnings.append(
                    f"Room {room_id} skipped because latest rent_end_date is before rent_start_date."
                )
                continue
            if rent_end_date >= self.today:
                skipped.append(f"Room {room_id} skipped because latest period has not ended.")
                continue

            new_start, new_end = self._next_period_dates(rent_start_date, rent_end_date)
            if new_end < new_start:
                warnings.append(
                    f"Room {room_id} skipped because the next rental period would be invalid."
                )
                continue
            if self._period_exists(payment_rows, room_id, new_start, new_end):
                skipped.append(
                    f"Room {room_id} skipped because {new_start} to {new_end} already exists."
                )
                continue

            new_row = {
                "room_id": room_id,
                "insert_after_row": latest["row_index"],
                "rent_start_date": new_start.isoformat(),
                "rent_end_date": new_end.isoformat(),
                "amount_due": _to_money(latest.get("amount_due")),
                "amount_paid": 0,
                "payment_date": None,
                "payment_status": self.calculate_payment_status(
                    room_status="Occupied",
                    rent_start_date=new_start,
                    payment_date=None,
                    rent_required=rent_required,
                ),
                "room_status": "Occupied",
                "payment_method": None,
                "tenant_name": _clean_text(latest.get("tenant_name")) or None,
                "tenant_ph": _clean_text(latest.get("tenant_ph")) or None,
                "ac": _normalize_yes_no(latest.get("ac")),
                "record_status": "Active",
                "source": "Auto rollover",
                "notes": "Assumed tenant continued",
            }
            rows_to_create.append(new_row)

        if preview:
            return OperationResponse(
                preview=True,
                success=True,
                message=f"Preview only. {len(rows_to_create)} rental period row(s) would be created.",
                rows_to_create=rows_to_create,
                skipped=skipped,
                warnings=warnings,
            )

        backup_path = self.create_backup()
        for row in sorted(rows_to_create, key=lambda item: item["insert_after_row"], reverse=True):
            new_row_index = self._insert_payment_row_after(
                payments_ws,
                payment_headers,
                after_row_index=int(row["insert_after_row"]),
                values=row,
            )
            self.write_audit_log(
                workbook,
                action_type="auto_rollover",
                room_id=_to_int(row["room_id"]),
                sheet_name="payments",
                row_index=new_row_index,
                old_value=None,
                new_value=row,
                user_message="Auto-created next rental period.",
                status="success",
            )
        self._save_workbook(workbook)
        return OperationResponse(
            preview=False,
            success=True,
            message=f"Created {len(rows_to_create)} rental period row(s).",
            backup_path=str(backup_path),
            rows_to_create=rows_to_create,
            skipped=skipped,
            warnings=warnings,
        )

    def record_next_period_payment(
        self,
        *,
        room_id: int,
        amount_paid: float,
        payment_date: str,
        payment_method: str,
        preview: bool = True,
        user_message: str | None = None,
    ) -> OperationResponse:
        paid_on = self._parse_request_date(payment_date, "payment_date")
        workbook, sheets = self._load_update_workbook()
        payments_ws = sheets["payments"]
        headers = self._headers(payments_ws)
        room = self._require_room(sheets["room_record"], room_id)
        rows = self._payment_rows(payments_ws, headers)
        latest = self._latest_active_payment_by_room(rows).get(room_id)
        if latest is None:
            raise ExcelDataError(f"No active payment row found for room {room_id}.")

        latest_start = self._parse_workbook_date(latest.get("rent_start_date"))
        latest_end = self._parse_workbook_date(latest.get("rent_end_date"))
        if latest_start is None or latest_end is None or latest_end < latest_start:
            raise ExcelDataError(f"Room {room_id}'s latest rental period is invalid.")

        next_start, next_end = self._next_period_dates(latest_start, latest_end)
        existing = next(
            (
                row for row in rows
                if row.get("room_id") == room_id
                and _normalize_record_status(row.get("record_status")) == "Active"
                and self._parse_workbook_date(row.get("rent_start_date")) == next_start
                and self._parse_workbook_date(row.get("rent_end_date")) == next_end
            ),
            None,
        )
        rent_required = _normalize_yes_no(room.get("rent_required"))
        status = self.calculate_payment_status(
            room_status="Occupied",
            rent_start_date=next_start,
            payment_date=paid_on,
            rent_required=rent_required,
        )

        if existing is not None:
            updates = {
                "amount_paid": amount_paid,
                "payment_date": paid_on,
                "payment_method": _normalize_payment_method(payment_method),
                "payment_status": status,
                "source": "API update",
            }
            changes = self._preview_row_updates(existing, updates, "payments")
            if preview:
                return self._preview_response(
                    f"Preview only. Payment for {next_start} to {next_end} would be updated.",
                    changes,
                )
            backup_path = self.create_backup()
            self._apply_row_updates(payments_ws, headers, existing["row_index"], updates)
            self.write_audit_log(
                workbook,
                action_type="record_next_period_payment",
                room_id=room_id,
                sheet_name="payments",
                row_index=existing["row_index"],
                old_value={change.column: change.old_value for change in changes},
                new_value={change.column: change.new_value for change in changes},
                user_message=user_message,
                status="success",
            )
            self._save_workbook(workbook)
            return self._saved_response("Next rental period payment updated.", backup_path, changes)

        new_row = {
            "room_id": room_id,
            "row_index": int(latest["row_index"]) + 1,
            "insert_after_row": latest["row_index"],
            "rent_start_date": next_start.isoformat(),
            "rent_end_date": next_end.isoformat(),
            "amount_due": _to_money(latest.get("amount_due")),
            "amount_paid": amount_paid,
            "payment_date": paid_on.isoformat(),
            "payment_status": status,
            "room_status": "Occupied",
            "payment_method": _normalize_payment_method(payment_method),
            "tenant_name": _clean_text(latest.get("tenant_name")) or None,
            "tenant_ph": _clean_text(latest.get("tenant_ph")) or None,
            "ac": _normalize_yes_no(latest.get("ac")),
            "record_status": "Active",
            "source": "API update",
            "notes": "Next rental period created when payment was recorded",
        }
        if preview:
            return OperationResponse(
                preview=True,
                success=True,
                message=f"Preview only. A paid rental period for {next_start} to {next_end} would be created.",
                rows_to_create=[new_row],
            )

        backup_path = self.create_backup()
        new_row_index = self._insert_payment_row_after(
            payments_ws,
            headers,
            after_row_index=int(latest["row_index"]),
            values=new_row,
        )
        self.write_audit_log(
            workbook,
            action_type="record_next_period_payment",
            room_id=room_id,
            sheet_name="payments",
            row_index=new_row_index,
            old_value=None,
            new_value=new_row,
            user_message=user_message,
            status="success",
        )
        self._save_workbook(workbook)
        return OperationResponse(
            preview=False,
            success=True,
            message="Next rental period payment created.",
            backup_path=str(backup_path),
            rows_to_create=[new_row],
        )

    def undo_last_change(
        self,
        *,
        preview: bool = True,
        backup_file: str | None = None,
    ) -> OperationResponse:
        backup_path = self._backup_for_restore(backup_file)
        description = {
            "backup_file": backup_path.name,
            "backup_created_at": datetime.fromtimestamp(
                backup_path.stat().st_mtime
            ).isoformat(timespec="seconds"),
            "workbook_to_restore": str(self.excel_path),
        }
        if preview:
            return OperationResponse(
                preview=True,
                success=True,
                message="Preview only. The workbook would be restored to its state before the last saved change.",
                rows_to_create=[description],
                warnings=["This restores the entire workbook, not only one cell."],
            )

        safety_backup = self.create_backup()
        try:
            shutil.copy2(backup_path, self.excel_path)
            restored = load_workbook(self.excel_path)
            self.write_audit_log(
                restored,
                action_type="undo_last_change",
                room_id=None,
                sheet_name="workbook",
                row_index=None,
                old_value={"safety_backup": str(safety_backup)},
                new_value={"restored_backup": str(backup_path)},
                user_message="Restored the workbook to its state before the last saved change.",
                status="success",
            )
            self._save_workbook(restored)
        except Exception as error:
            shutil.copy2(safety_backup, self.excel_path)
            raise ExcelDataError(
                f"Undo failed. The current workbook was restored from its safety backup: {error}"
            ) from error
        return OperationResponse(
            preview=False,
            success=True,
            message=f"Last change undone using {backup_path.name}.",
            backup_path=str(safety_backup),
            rows_to_create=[description],
        )

    def _backup_for_restore(self, backup_file: str | None = None) -> Path:
        if backup_file:
            candidate = (self.backup_dir / Path(backup_file).name).resolve()
            if candidate.parent != self.backup_dir.resolve() or not candidate.exists():
                raise ExcelDataError(f"Backup {backup_file} is not available.")
            return candidate
        backups = sorted(
            self.backup_dir.glob("rental_backup_*.xlsx"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        if not backups:
            raise ExcelDataError("No rental backup is available to undo.")
        return backups[0]

    def update_current_payment(
        self,
        *,
        room_id: int,
        amount_paid: float,
        payment_date: str,
        payment_method: str,
        rent_start_date: str | None = None,
        rent_end_date: str | None = None,
        preview: bool = True,
        user_message: str | None = None,
    ) -> OperationResponse:
        payment_date_value = self._parse_request_date(payment_date, "payment_date")
        start_value = self._parse_request_date(rent_start_date, "rent_start_date") if rent_start_date else None
        end_value = self._parse_request_date(rent_end_date, "rent_end_date") if rent_end_date else None
        workbook, sheets = self._load_update_workbook()
        payments_ws = sheets["payments"]
        headers = self._headers(payments_ws)
        room = self._require_room(sheets["room_record"], room_id)
        row = self._find_target_payment(payments_ws, headers, room_id, start_value, end_value)

        new_status = self.calculate_payment_status(
            room_status=_normalize_room_status(row.get("room_status")),
            rent_start_date=self._parse_workbook_date(row.get("rent_start_date")),
            payment_date=payment_date_value,
            rent_required=_normalize_yes_no(room.get("rent_required")),
        )
        updates = {
            "amount_paid": amount_paid,
            "payment_date": payment_date_value,
            "payment_method": _normalize_payment_method(payment_method),
            "payment_status": new_status,
            "source": "API update",
        }
        changes = self._preview_row_updates(row, updates, "payments")
        if preview:
            return self._preview_response("Preview only. Current payment would be updated.", changes)

        backup_path = self.create_backup()
        self._apply_row_updates(payments_ws, headers, row["row_index"], updates)
        self.write_audit_log(
            workbook,
            action_type="update_current_payment",
            room_id=room_id,
            sheet_name="payments",
            row_index=row["row_index"],
            old_value={change.column: change.old_value for change in changes},
            new_value={change.column: change.new_value for change in changes},
            user_message=user_message,
            status="success",
        )
        self._save_workbook(workbook)
        return self._saved_response("Current payment updated.", backup_path, changes)

    def move_out_tenant(
        self,
        *,
        room_id: int,
        move_out_date: str,
        preview: bool = True,
        user_message: str | None = None,
    ) -> OperationResponse:
        move_out = self._parse_request_date(move_out_date, "move_out_date")
        workbook, sheets = self._load_update_workbook()
        room_ws = sheets["room_record"]
        payments_ws = sheets["payments"]
        room_headers = self._headers(room_ws)
        payment_headers = self._headers(payments_ws)
        room = self._require_room(room_ws, room_id)
        target = self._find_target_payment(payments_ws, payment_headers, room_id, None, None)
        payment_rows = self._payment_rows(payments_ws, payment_headers)
        target_is_paid = _to_money(target.get("amount_paid")) > 0 or self._parse_workbook_date(target.get("payment_date")) is not None
        warnings: list[str] = []

        changes: list[ChangePreview] = []
        target_updates: dict[str, Any] = {}
        if target_is_paid:
            warnings.append(
                "Latest payment row already has payment. It will keep its room/payment status and only receive a move-out note."
            )
            target_updates = {
                "notes": self._join_notes(target.get("notes"), f"Tenant moved out on {move_out}."),
                "source": "API update",
            }
            changes.extend(self._preview_row_updates(target, target_updates, "payments"))
        else:
            target_updates = {
                "room_status": "Empty",
                "payment_status": "N/A",
                "notes": self._join_notes(target.get("notes"), f"Tenant moved out on {move_out}."),
                "source": "API update",
            }
            changes.extend(self._preview_row_updates(target, target_updates, "payments"))

        room_changes = self._preview_row_updates(
            room,
            {"current_occupants": 0},
            "room_record",
        )
        changes.extend(room_changes)

        future_rows = [
            row for row in payment_rows
            if row.get("room_id") == room_id
            and _normalize_record_status(row.get("record_status")) == "Active"
            and self._parse_workbook_date(row.get("rent_start_date")) is not None
            and self._parse_workbook_date(row.get("rent_start_date")) > move_out
        ]
        for row in future_rows:
            changes.extend(
                self._preview_row_updates(
                    row,
                    {
                        "record_status": "Cancelled",
                        "notes": self._join_notes(row.get("notes"), f"Cancelled because tenant moved out on {move_out}."),
                    },
                    "payments",
                )
            )

        if preview:
            response = self._preview_response("Preview only. Tenant move-out would be applied.", changes)
            response.warnings = warnings
            return response

        backup_path = self.create_backup()
        if target_updates:
            self._apply_row_updates(payments_ws, payment_headers, target["row_index"], target_updates)
        self._apply_row_updates(room_ws, room_headers, room["row_index"], {
            "current_occupants": 0,
        })
        for row in future_rows:
            self._apply_row_updates(payments_ws, payment_headers, row["row_index"], {
                "record_status": "Cancelled",
                "notes": self._join_notes(row.get("notes"), f"Cancelled because tenant moved out on {move_out}."),
            })
        self.write_audit_log(
            workbook,
            action_type="move_out_tenant",
            room_id=room_id,
            sheet_name="payments",
            row_index=target["row_index"],
            old_value={change.column: change.old_value for change in changes},
            new_value={change.column: change.new_value for change in changes},
            user_message=user_message,
            status="success",
        )
        self._save_workbook(workbook)
        response = self._saved_response("Tenant move-out saved.", backup_path, changes)
        response.warnings = warnings
        return response

    def move_in_tenant(
        self,
        *,
        room_id: int,
        tenant_name: str,
        tenant_ph: str | None,
        rent_start_date: str,
        rent_end_date: str,
        amount_due: float,
        current_occupants: int,
        ac: str,
        amount_paid: float = 0,
        payment_date: str | None = None,
        payment_method: str | None = None,
        preview: bool = True,
        user_message: str | None = None,
    ) -> OperationResponse:
        if not _clean_text(tenant_name):
            raise ExcelDataError("tenant_name is required for move-in.")
        start = self._parse_request_date(rent_start_date, "rent_start_date")
        end = self._parse_request_date(rent_end_date, "rent_end_date")
        paid_on = self._parse_request_date(payment_date, "payment_date") if payment_date else None
        ac_value = _normalize_yes_no(ac)
        if ac_value not in {"Y", "N"}:
            raise ExcelDataError("ac must be Y or N.")

        workbook, sheets = self._load_update_workbook()
        room_ws = sheets["room_record"]
        payments_ws = sheets["payments"]
        room_headers = self._headers(room_ws)
        payment_headers = self._headers(payments_ws)
        room = self._require_room(room_ws, room_id)
        warnings = [] if tenant_ph else ["tenant_ph was not provided."]
        insert_after_row = self._latest_payment_row_for_room(payments_ws, payment_headers, room_id)

        payment_row = {
            "room_id": room_id,
            "insert_after_row": insert_after_row,
            "rent_start_date": start.isoformat(),
            "rent_end_date": end.isoformat(),
            "amount_due": amount_due,
            "amount_paid": amount_paid,
            "payment_date": paid_on.isoformat() if paid_on else None,
            "payment_status": self.calculate_payment_status(
                room_status="Occupied",
                rent_start_date=start,
                payment_date=paid_on,
                rent_required="Y",
            ),
            "room_status": "Occupied",
            "payment_method": _normalize_payment_method(payment_method),
            "tenant_name": tenant_name,
            "tenant_ph": tenant_ph,
            "ac": ac_value,
            "record_status": "Active",
            "source": "API update",
            "notes": "New tenant moved in",
        }
        room_updates = {
            "current_occupants": current_occupants,
            "ac": ac_value,
            "usual_price": amount_due,
            "rent_required": "Y",
        }
        changes = self._preview_row_updates(room, room_updates, "room_record")

        if preview:
            return OperationResponse(
                preview=True,
                success=True,
                message="Preview only. New tenant row would be added.",
                rows_to_create=[payment_row],
                changes=changes,
                warnings=warnings,
            )

        backup_path = self.create_backup()
        new_row_index = self._insert_payment_row_after(
            payments_ws,
            payment_headers,
            after_row_index=insert_after_row,
            values=payment_row,
        )
        self._apply_row_updates(room_ws, room_headers, room["row_index"], room_updates)
        self.write_audit_log(
            workbook,
            action_type="move_in_tenant",
            room_id=room_id,
            sheet_name="payments",
            row_index=new_row_index,
            old_value=None,
            new_value=payment_row,
            user_message=user_message,
            status="success",
        )
        self._save_workbook(workbook)
        return OperationResponse(
            preview=False,
            success=True,
            message="New tenant saved.",
            backup_path=str(backup_path),
            rows_to_create=[payment_row],
            changes=changes,
            warnings=warnings,
        )

    def update_room_rent(
        self,
        *,
        room_id: int,
        new_amount_due: float,
        effective_start_date: str | None = None,
        ac: str | None = None,
        current_occupants: int | None = None,
        preview: bool = True,
        user_message: str | None = None,
    ) -> OperationResponse:
        effective = self._parse_request_date(effective_start_date, "effective_start_date") if effective_start_date else None
        ac_value = _normalize_yes_no(ac) if ac is not None else None
        if ac is not None and ac_value not in {"Y", "N"}:
            raise ExcelDataError("ac must be Y or N.")
        if current_occupants is not None and current_occupants < 0:
            raise ExcelDataError("current_occupants cannot be negative.")
        workbook, sheets = self._load_update_workbook()
        room_ws = sheets["room_record"]
        payments_ws = sheets["payments"]
        room_headers = self._headers(room_ws)
        payment_headers = self._headers(payments_ws)
        room = self._require_room(room_ws, room_id)
        payment_targets = self._rent_update_targets(payments_ws, payment_headers, room_id, effective)
        warnings = []
        future_payment_row = None
        rent_note = self._rent_change_note(
            new_amount_due=new_amount_due,
            effective_start_date=effective,
            reason=user_message,
        )
        if not payment_targets and effective is not None:
            future_payment_row = self._build_future_rent_change_row(
                payments_ws,
                payment_headers,
                room=room,
                room_id=room_id,
                effective_start_date=effective,
                new_amount_due=new_amount_due,
                ac=ac_value,
                note=rent_note,
            )
            if future_payment_row is None:
                warnings.append("No active occupied payment row found to copy into a future rent-change row.")
        elif not payment_targets:
            warnings.append("No current/future active payment rows matched; only room_record would be updated.")

        room_updates: dict[str, Any] = {"usual_price": new_amount_due}
        if ac_value is not None:
            room_updates["ac"] = ac_value
        if current_occupants is not None:
            room_updates["current_occupants"] = current_occupants

        room_changes = self._preview_row_updates(room, room_updates, "room_record")
        changes = room_changes[:]
        for target in payment_targets:
            payment_updates: dict[str, Any] = {
                "amount_due": new_amount_due,
                "source": "API update",
                "notes": self._join_notes(target.get("notes"), rent_note),
            }
            if ac_value is not None:
                payment_updates["ac"] = ac_value
            changes.extend(
                self._preview_row_updates(
                    target,
                    payment_updates,
                    "payments",
                )
            )
        if preview:
            return OperationResponse(
                preview=True,
                success=True,
                message="Preview only. Room rent would be updated.",
                rows_to_create=[future_payment_row] if future_payment_row else [],
                changes=changes,
                warnings=warnings,
            )

        backup_path = self.create_backup()
        new_row_index = None
        if future_payment_row:
            new_row_index = self._insert_payment_row_after(
                payments_ws,
                payment_headers,
                after_row_index=int(future_payment_row["insert_after_row"]),
                values=future_payment_row,
            )
        self._apply_row_updates(room_ws, room_headers, room["row_index"], room_updates)
        for target in payment_targets:
            payment_updates = {
                "amount_due": new_amount_due,
                "source": "API update",
                "notes": self._join_notes(target.get("notes"), rent_note),
            }
            if ac_value is not None:
                payment_updates["ac"] = ac_value
            self._apply_row_updates(payments_ws, payment_headers, target["row_index"], payment_updates)
        self.write_audit_log(
            workbook,
            action_type="update_room_rent",
            room_id=room_id,
            sheet_name="room_record/payments",
            row_index=new_row_index or room["row_index"],
            old_value={change.column: change.old_value for change in changes},
            new_value=future_payment_row or {change.column: change.new_value for change in changes},
            user_message=user_message,
            status="success",
        )
        self._save_workbook(workbook)
        response = OperationResponse(
            preview=False,
            success=True,
            message="Room rent updated.",
            backup_path=str(backup_path),
            rows_to_create=[future_payment_row] if future_payment_row else [],
            changes=changes,
            warnings=warnings,
        )
        return response

    def get_audit_log(self, limit: int = 50) -> list[dict[str, Any]]:
        workbook, sheets = self._load_update_workbook()
        audit_ws = sheets["audit_log"]
        headers = self._headers(audit_ws)
        rows = [self._row_to_dict(audit_ws, headers, row_index) for row_index in range(2, audit_ws.max_row + 1)]
        return rows[-limit:]

    def create_backup(self) -> Path:
        if not self.excel_path.exists():
            raise ExcelDataError(f"Excel file not found: {self.excel_path}")
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = self.backup_dir / f"rental_backup_{timestamp}.xlsx"
        counter = 1
        while backup_path.exists():
            backup_path = self.backup_dir / f"rental_backup_{timestamp}_{counter}.xlsx"
            counter += 1
        try:
            shutil.copy2(self.excel_path, backup_path)
        except OSError as error:
            raise ExcelDataError(f"Backup failed. Excel was not updated: {error}") from error
        return backup_path

    def write_audit_log(
        self,
        workbook: Any,
        *,
        action_type: str,
        room_id: int | None,
        sheet_name: str,
        row_index: int | None,
        old_value: Any,
        new_value: Any,
        user_message: str | None,
        status: str,
    ) -> None:
        audit_ws = self._ensure_audit_sheet(workbook)
        audit_ws.append([
            datetime.now().isoformat(timespec="seconds"),
            action_type,
            room_id,
            sheet_name,
            row_index,
            self._json_value(old_value),
            self._json_value(new_value),
            user_message,
            status,
        ])

    def calculate_payment_status(
        self,
        *,
        room_status: str | None,
        rent_start_date: date | None,
        payment_date: date | None,
        rent_required: str | None,
    ) -> str:
        if _normalize_yes_no(rent_required) == "N":
            return "N/A"
        if _normalize_room_status(room_status) == "Empty":
            return "N/A"
        if rent_start_date is None:
            return "Unknown"
        if payment_date is not None:
            return "Late" if payment_date > rent_start_date + timedelta(days=7) else "Paid"
        return "Unpaid" if self.today > rent_start_date + timedelta(days=7) else "Pending"

    def _load_update_workbook(self) -> tuple[Any, dict[str, Worksheet]]:
        if not self.excel_path.exists():
            raise ExcelDataError(f"Excel file not found: {self.excel_path}")
        workbook = load_workbook(self.excel_path)
        missing_sheets = sorted({"room_record", "payments"} - set(workbook.sheetnames))
        if missing_sheets:
            raise ExcelDataError(
                f"Missing required sheet(s): {', '.join(missing_sheets)}",
                available_sheets=workbook.sheetnames,
            )
        payments_ws = workbook["payments"]
        self._ensure_columns(payments_ws, PAYMENT_EXTRA_COLUMNS)
        audit_ws = self._ensure_audit_sheet(workbook)
        return workbook, {
            "room_record": workbook["room_record"],
            "payments": payments_ws,
            "audit_log": audit_ws,
        }

    def _ensure_audit_sheet(self, workbook: Any) -> Worksheet:
        if "audit_log" not in workbook.sheetnames:
            audit_ws = workbook.create_sheet("audit_log")
            audit_ws.append(AUDIT_COLUMNS)
            return audit_ws
        audit_ws = workbook["audit_log"]
        self._ensure_columns(audit_ws, AUDIT_COLUMNS)
        return audit_ws

    def _ensure_columns(self, worksheet: Worksheet, column_names: list[str]) -> None:
        headers = self._headers(worksheet)
        for column_name in column_names:
            normalized = _normalize_column(column_name)
            if normalized not in headers:
                worksheet.cell(row=1, column=worksheet.max_column + 1, value=column_name)
                headers[normalized] = worksheet.max_column

    def _headers(self, worksheet: Worksheet) -> dict[str, int]:
        headers: dict[str, int] = {}
        for cell in worksheet[1]:
            if cell.value is None:
                continue
            normalized = _normalize_column(cell.value)
            if normalized == "ac_portable_ac":
                normalized = "ac"
            headers[normalized] = cell.column
        return headers

    def _row_to_dict(self, worksheet: Worksheet, headers: dict[str, int], row_index: int) -> dict[str, Any]:
        row = {"row_index": row_index}
        for name, column_index in headers.items():
            row[name] = worksheet.cell(row=row_index, column=column_index).value
        return row

    def _room_by_id(self, room_ws: Worksheet) -> dict[int, dict[str, Any]]:
        headers = self._headers(room_ws)
        rows: dict[int, dict[str, Any]] = {}
        for row_index in range(2, room_ws.max_row + 1):
            row = self._row_to_dict(room_ws, headers, row_index)
            room_id = _to_int(row.get("room_id"))
            if room_id is not None:
                row["room_id"] = room_id
                rows[room_id] = row
        return rows

    def _require_room(self, room_ws: Worksheet, room_id: int) -> dict[str, Any]:
        rooms = self._room_by_id(room_ws)
        if room_id not in rooms:
            available_room_ids = sorted(rooms)
            raise ExcelDataError(
                f"Room {room_id} was not found. Available room IDs: {available_room_ids}"
            )
        return rooms[room_id]

    def _payment_rows(self, payments_ws: Worksheet, headers: dict[str, int]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row_index in range(2, payments_ws.max_row + 1):
            row = self._row_to_dict(payments_ws, headers, row_index)
            room_id = _to_int(row.get("room_id"))
            if room_id is None:
                continue
            row["room_id"] = room_id
            rows.append(row)
        return rows

    def _latest_active_payment_by_room(self, rows: list[dict[str, Any]]) -> dict[int, dict[str, Any]]:
        latest: dict[int, dict[str, Any]] = {}
        for row in rows:
            if _normalize_record_status(row.get("record_status")) != "Active":
                continue
            room_id = _to_int(row.get("room_id"))
            if room_id is None:
                continue
            current = latest.get(room_id)
            if current is None or self._row_sort_key(row) > self._row_sort_key(current):
                latest[room_id] = row
        return latest

    def _latest_payment_row_for_room(
        self,
        payments_ws: Worksheet,
        headers: dict[str, int],
        room_id: int,
    ) -> int:
        room_rows = [
            row for row in self._payment_rows(payments_ws, headers)
            if row.get("room_id") == room_id
        ]
        if not room_rows:
            return self._last_data_row(payments_ws, headers)
        return sorted(room_rows, key=self._row_sort_key)[-1]["row_index"]

    def _find_target_payment(
        self,
        payments_ws: Worksheet,
        headers: dict[str, int],
        room_id: int,
        rent_start_date: date | None,
        rent_end_date: date | None,
    ) -> dict[str, Any]:
        rows = [
            row for row in self._payment_rows(payments_ws, headers)
            if row.get("room_id") == room_id and _normalize_record_status(row.get("record_status")) == "Active"
        ]
        if rent_start_date:
            rows = [row for row in rows if self._parse_workbook_date(row.get("rent_start_date")) == rent_start_date]
        if rent_end_date:
            rows = [row for row in rows if self._parse_workbook_date(row.get("rent_end_date")) == rent_end_date]
        if not rows:
            raise ExcelDataError(f"No active payment row found for room {room_id}.")
        if rent_start_date is None and rent_end_date is None:
            latest = self._latest_active_payment_by_room(rows).get(room_id)
            if latest is None:
                raise ExcelDataError(f"No active payment row found for room {room_id}.")
            return latest
        if len(rows) > 1:
            matching = [
                {
                    "row_index": row["row_index"],
                    "rent_start_date": self._date_or_none(row.get("rent_start_date")),
                    "rent_end_date": self._date_or_none(row.get("rent_end_date")),
                    "tenant_name": row.get("tenant_name"),
                }
                for row in rows
            ]
            raise ExcelDataError(
                f"Multiple active payment rows found for room {room_id}. "
                f"Please specify the rent period. Matching rows: {matching}"
            )
        return rows[0]

    def _rent_update_targets(
        self,
        payments_ws: Worksheet,
        headers: dict[str, int],
        room_id: int,
        effective_start_date: date | None,
    ) -> list[dict[str, Any]]:
        rows = [
            row for row in self._payment_rows(payments_ws, headers)
            if row.get("room_id") == room_id and _normalize_record_status(row.get("record_status")) == "Active"
        ]
        if effective_start_date is None:
            latest = self._latest_active_payment_by_room(rows).get(room_id)
            return [latest] if latest else []

        targets = []
        for row in rows:
            start = self._parse_workbook_date(row.get("rent_start_date"))
            if start and start >= effective_start_date:
                targets.append(row)
        return sorted(targets, key=self._row_sort_key)

    def _build_future_rent_change_row(
        self,
        payments_ws: Worksheet,
        headers: dict[str, int],
        *,
        room: dict[str, Any],
        room_id: int,
        effective_start_date: date,
        new_amount_due: float,
        ac: str | None,
        note: str,
    ) -> dict[str, Any] | None:
        rows = [
            row for row in self._payment_rows(payments_ws, headers)
            if row.get("room_id") == room_id
            and _normalize_record_status(row.get("record_status")) == "Active"
        ]
        latest = self._latest_active_payment_by_room(rows).get(room_id)
        if latest is None:
            return None
        if _normalize_room_status(latest.get("room_status")) != "Occupied":
            return None
        if _normalize_yes_no(room.get("rent_required")) == "N":
            return None

        latest_start = self._parse_workbook_date(latest.get("rent_start_date"))
        latest_end = self._parse_workbook_date(latest.get("rent_end_date"))
        if latest_start is None or latest_end is None or latest_end < latest_start:
            return None

        effective_end = self._next_period_end_from_pattern(
            latest_start,
            latest_end,
            effective_start_date,
        )
        if self._period_exists(rows, room_id, effective_start_date, effective_end):
            return None

        return {
            "room_id": room_id,
            "insert_after_row": latest["row_index"],
            "rent_start_date": effective_start_date.isoformat(),
            "rent_end_date": effective_end.isoformat(),
            "amount_due": new_amount_due,
            "amount_paid": 0,
            "payment_date": None,
            "payment_status": self.calculate_payment_status(
                room_status="Occupied",
                rent_start_date=effective_start_date,
                payment_date=None,
                rent_required=_normalize_yes_no(room.get("rent_required")),
            ),
            "room_status": "Occupied",
            "payment_method": None,
            "tenant_name": _clean_text(latest.get("tenant_name")) or None,
            "tenant_ph": _clean_text(latest.get("tenant_ph")) or None,
            "ac": ac or _normalize_yes_no(latest.get("ac")),
            "record_status": "Active",
            "source": "API update",
            "notes": note,
        }

    def _period_exists(
        self,
        rows: list[dict[str, Any]],
        room_id: int,
        rent_start_date: date,
        rent_end_date: date,
    ) -> bool:
        for row in rows:
            if row.get("room_id") != room_id:
                continue
            if _normalize_record_status(row.get("record_status")) != "Active":
                continue
            if self._parse_workbook_date(row.get("rent_start_date")) == rent_start_date and self._parse_workbook_date(row.get("rent_end_date")) == rent_end_date:
                return True
        return False

    def _append_payment_row(
        self, payments_ws: Worksheet, headers: dict[str, int], values: dict[str, Any]
    ) -> None:
        row_index = self._last_data_row(payments_ws, headers) + 1
        self._write_payment_row(payments_ws, headers, row_index, values)

    def _insert_payment_row_after(
        self,
        payments_ws: Worksheet,
        headers: dict[str, int],
        *,
        after_row_index: int,
        values: dict[str, Any],
    ) -> int:
        row_index = after_row_index + 1
        payments_ws.insert_rows(row_index)
        self._copy_row_style(payments_ws, source_row=after_row_index, target_row=row_index)
        self._write_payment_row(payments_ws, headers, row_index, values)
        return row_index

    def _write_payment_row(
        self,
        payments_ws: Worksheet,
        headers: dict[str, int],
        row_index: int,
        values: dict[str, Any],
    ) -> None:
        for name, value in values.items():
            if name == "insert_after_row":
                continue
            column_index = headers.get(_normalize_column(name))
            if column_index:
                payments_ws.cell(row=row_index, column=column_index, value=self._excel_value(value))

    def _copy_row_style(self, worksheet: Worksheet, *, source_row: int, target_row: int) -> None:
        for column_index in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=source_row, column=column_index)
            target_cell = worksheet.cell(row=target_row, column=column_index)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)

    def _last_data_row(self, worksheet: Worksheet, headers: dict[str, int]) -> int:
        key_columns = [
            headers[column]
            for column in ("room_id", "rent_start_date", "rent_end_date", "tenant_name")
            if column in headers
        ]
        if not key_columns:
            key_columns = list(headers.values())

        for row_index in range(worksheet.max_row, 1, -1):
            has_data = any(
                worksheet.cell(row=row_index, column=column_index).value not in (None, "")
                for column_index in key_columns
            )
            if has_data:
                return row_index
        return 1

    def _preview_row_updates(
        self,
        row: dict[str, Any],
        updates: dict[str, Any],
        sheet_name: str,
    ) -> list[ChangePreview]:
        changes: list[ChangePreview] = []
        for column, new_value in updates.items():
            old_value = row.get(column)
            clean_new = self._excel_value(new_value)
            if self._json_value(old_value) == self._json_value(clean_new):
                continue
            changes.append(
                ChangePreview(
                    sheet_name=sheet_name,
                    row_index=row.get("row_index"),
                    column=column,
                    old_value=self._api_value(old_value),
                    new_value=self._api_value(clean_new),
                )
            )
        return changes

    def _apply_row_updates(
        self,
        worksheet: Worksheet,
        headers: dict[str, int],
        row_index: int,
        updates: dict[str, Any],
    ) -> None:
        for column, value in updates.items():
            column_index = headers.get(_normalize_column(column))
            if column_index is None:
                raise ExcelDataError(f"Column {column} is missing from {worksheet.title}.")
            worksheet.cell(row=row_index, column=column_index, value=self._excel_value(value))

    def _preview_response(self, message: str, changes: list[ChangePreview]) -> OperationResponse:
        return OperationResponse(preview=True, success=True, message=message, changes=changes)

    def _saved_response(
        self,
        message: str,
        backup_path: Path,
        changes: list[ChangePreview],
    ) -> OperationResponse:
        return OperationResponse(
            preview=False,
            success=True,
            message=message,
            backup_path=str(backup_path),
            changes=changes,
        )

    def _save_workbook(self, workbook: Any) -> None:
        try:
            workbook.save(self.excel_path)
        except OSError as error:
            raise ExcelDataError(f"Saving failed. Backup was kept: {error}") from error

    def _next_period_dates(self, rent_start_date: date, rent_end_date: date) -> tuple[date, date]:
        if rent_start_date.day == rent_end_date.day:
            return rent_end_date, self._add_one_month(rent_end_date)
        if rent_start_date.day == 1 and rent_end_date.day == self._last_day(rent_end_date):
            new_start = rent_end_date + timedelta(days=1)
            return new_start, date(new_start.year, new_start.month, self._last_day(new_start))
        period_length = rent_end_date - rent_start_date
        new_start = rent_end_date + timedelta(days=1)
        return new_start, new_start + period_length

    def _next_period_end_from_pattern(
        self,
        previous_start: date,
        previous_end: date,
        new_start: date,
    ) -> date:
        if previous_start.day == previous_end.day:
            return self._add_one_month(new_start)
        if previous_start.day == 1 and previous_end.day == self._last_day(previous_end):
            return date(new_start.year, new_start.month, self._last_day(new_start))
        return new_start + (previous_end - previous_start)

    def _add_one_month(self, value: date) -> date:
        month = value.month + 1
        year = value.year
        if month == 13:
            month = 1
            year += 1
        return date(year, month, min(value.day, monthrange(year, month)[1]))

    def _last_day(self, value: date) -> int:
        return monthrange(value.year, value.month)[1]

    def _row_sort_key(self, row: dict[str, Any]) -> tuple[str, str, int]:
        return (
            self._date_or_none(row.get("rent_start_date")) or "",
            self._date_or_none(row.get("rent_end_date")) or "",
            int(row.get("row_index") or 0),
        )

    def _parse_request_date(self, value: str | None, field_name: str) -> date:
        if not value:
            raise ExcelDataError(f"{field_name} is required.")
        try:
            return date.fromisoformat(value)
        except ValueError as error:
            raise ExcelDataError(f"{field_name} must be a valid YYYY-MM-DD date.") from error

    def _parse_workbook_date(self, value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                return from_excel(value).date()
            except (TypeError, ValueError):
                return None
        text = _clean_text(value)
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None

    def _date_or_none(self, value: Any) -> str | None:
        parsed = self._parse_workbook_date(value)
        return parsed.isoformat() if parsed else None

    def _excel_value(self, value: Any) -> Any:
        if isinstance(value, date) and not isinstance(value, datetime):
            return value.isoformat()
        return value

    def _api_value(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return value

    def _json_value(self, value: Any) -> str | None:
        if value is None:
            return None
        return json.dumps(self._api_value(value), default=str, ensure_ascii=True, sort_keys=True)

    def _rent_change_note(
        self,
        *,
        new_amount_due: float,
        effective_start_date: date | None,
        reason: str | None,
    ) -> str:
        note = f"Rent changed to {float(new_amount_due)}"
        if effective_start_date is not None:
            note += f" effective {effective_start_date.isoformat()}"
        clean_reason = _clean_text(reason)
        if clean_reason:
            note += f" due to {clean_reason}"
        return f"{note}."

    def _join_notes(self, existing: Any, new_note: str) -> str:
        existing_text = _clean_text(existing)
        if existing_text == new_note or f"; {new_note}" in existing_text:
            return existing_text
        return f"{existing_text}; {new_note}" if existing_text else new_note
